#!/usr/bin/env python3
"""Converte o "Relatório Estrutura Nível" do ERP para a aba ESTRUTURA.

    python3 converter-estrutura.py ESTRUTURA_NIVEL.xlsx [saida.csv] [--ignorar 600,611]

O relatório do ERP é feito para ler no papel, não para ser lido por programa:
o pai vem numa linha solta e os filhos abaixo dele, sem repetir a chave. Este
script achata isso em CODIGO / PECA / QTD / DESCRICAO — uma linha por peça,
que é o formato que a tela de faltas lê.

Serve para o export de 6 volumes ou para o catálogo inteiro; o formato é o
mesmo. Reprocessar não faz mal: a saída é determinística.

GRUPOS IGNORADOS (os 3 primeiros dígitos da peça)
A estrutura do ERP mistura peça produzida com insumo medido. Cola e chapa
entram fracionados — 0,04 e 0,3845 por volume — porque são consumo, não peça
contada; ninguém reporta "falta 0,04 de cola" para fechar um lote. Ficam de
fora por padrão. Embalagem (607) e acessórios (603) CONTINUAM na lista: caixa
ou corrediça faltando trava o volume igual, e o operador precisa poder dizer.
Use --ignorar para mudar, ou --ignorar "" para não filtrar nada.
"""
import sys, re, csv

IGNORAR_PADRAO = {"600", "611"}   # 600 chapa HDF · 611 cola

try:
    import openpyxl
except ImportError:
    sys.exit("falta o openpyxl:  pip install openpyxl")

# Colunas do relatório (descobertas no export real; o cabeçalho impresso está
# desalinhado dos dados, então vale a posição, não o título):
#   A=Nível  E=código do pai  F=código do filho  G=descrição
#   H=Qtd. Est.  J=Qtd. Eng.  M=fase  N=peso
COL_NIVEL, COL_PAI, COL_FILHO, COL_DESC = 1, 5, 6, 7
COL_QTD_EST, COL_QTD_ENG = 8, 10

cod = lambda v: re.sub(r"[^0-9A-Z]", "", str(v).upper()) if v not in (None, "") else ""


def converter(caminho, ignorar=IGNORAR_PADRAO):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb[wb.sheetnames[0]]
    pai, out, avisos = None, [], []
    vistos = set()
    ignoradas = 0

    for r in range(1, ws.max_row + 1):
        e = cod(ws.cell(r, COL_PAI).value)
        nivel = str(ws.cell(r, COL_NIVEL).value or "").strip()
        f = cod(ws.cell(r, COL_FILHO).value)

        # linha de pai: traz código em E e não tem nível preenchido.
        # exige 9 dígitos para não confundir com as linhas de título do
        # relatório ("RELATÓRIO ESTRUTURA NÍVEL ÚNICO" caía aqui)
        if re.fullmatch(r"\d{9}", e) and not nivel:
            pai = e
            continue

        if not pai or nivel != "1" or not re.fullmatch(r"\d{9}", f):
            continue

        if f[:3] in ignorar:          # insumo medido, não peça contada
            ignoradas += 1
            continue

        est = ws.cell(r, COL_QTD_EST).value
        eng = ws.cell(r, COL_QTD_ENG).value
        if est is not None and eng is not None and est != eng:
            avisos.append(f"{pai} -> {f}: Qtd.Est={est} difere de Qtd.Eng={eng}")

        qtd = est if est is not None else eng
        if qtd in (None, ""):
            continue

        chave = (pai, f)
        if chave in vistos:            # mesma peça repetida no relatório
            continue
        vistos.add(chave)

        # inteiro sai sem casa decimal: "1" lê melhor que "1.0" na planilha
        q = float(qtd)
        out.append([pai, f, int(q) if q == int(q) else q,
                    str(ws.cell(r, COL_DESC).value or "").strip()])

    return out, avisos, ignoradas


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    argv = sys.argv[1:]
    ignorar = IGNORAR_PADRAO
    if "--ignorar" in argv:
        i = argv.index("--ignorar")
        ignorar = {g.strip() for g in argv[i + 1].split(",") if g.strip()}
        del argv[i:i + 2]
    entrada = argv[0]
    saida = argv[1] if len(argv) > 1 else "ESTRUTURA.csv"

    linhas, avisos, ignoradas = converter(entrada, ignorar)
    if not linhas:
        sys.exit("nenhum par volume-peça encontrado — confira se é o "
                 "Relatório Estrutura Nível")

    # utf-8-sig: sem o BOM o Excel abre os acentos errados
    with open(saida, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(["CODIGO", "PECA", "QTD", "DESCRICAO"])
        w.writerows(linhas)

    volumes = len({l[0] for l in linhas})
    pecas = len({l[1] for l in linhas})
    print(f"{saida}: {len(linhas)} pares · {volumes} volumes · {pecas} peças "
          f"distintas · {len(linhas)/volumes:.1f} por volume")
    if ignoradas:
        print(f"  {ignoradas} linhas ignoradas (grupos {sorted(ignorar)})")

    # peça compartilhada é o que faz o relatório "destrava N volumes" valer:
    # uma corrediça parada pode segurar dezenas de lotes
    import collections
    uso = collections.Counter(l[1] for l in linhas)
    comp = sum(1 for v in uso.values() if v > 1)
    if comp:
        print(f"  {comp} peças usadas em mais de um volume "
              f"(máx. {max(uso.values())})")

    if avisos:
        print(f"\n{len(avisos)} linha(s) com Qtd.Est != Qtd.Eng "
              f"(usei a Est.):")
        for a in avisos[:10]:
            print("  " + a)


if __name__ == "__main__":
    main()
